import openpyxl
from openpyxl.styles import Font, PatternFill
from datetime import date, timedelta
import os

def create_case_study_excel(file_path="studi_kasus_yayasan_2025.xlsx"):
    wb = openpyxl.Workbook()

    # --- Sheet 1: Bagan Akun Baru ---
    ws_coa = wb.active
    ws_coa.title = "Bagan Akun Baru"
    coa_headers = ["Kode", "Nama Akun", "Tipe", "Kategori"]
    ws_coa.append(coa_headers)

    # Apply header style
    for col_idx, header in enumerate(coa_headers, 1):
        ws_coa.cell(row=1, column=col_idx).font = Font(bold=True)
        ws_coa.cell(row=1, column=col_idx).fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")

    new_accounts = [
        ("1003", "Kas", "Asset", "Without Restriction"), # Mengganti Kas Kecil dengan Kas
        ("1004", "Bank Mandiri", "Asset", "Without Restriction"),
        ("3001", "Aset Neto Tidak Terikat (Surplus/Defisit)", "Asset Net", "Without Restriction"),
        ("4001", "Pendapatan Donasi", "Revenue", "Without Restriction"),
        ("5001", "Beban Sewa Kantor", "Expense", "Without Restriction"),
        ("5002", "Beban Gaji Karyawan (Admin)", "Expense", "Without Restriction"),
        ("5003", "Beban Operasional Lainnya", "Expense", "Without Restriction"),
        ("1501", "Peralatan Kantor", "Asset", "Without Restriction"),
        ("1502", "Akumulasi Penyusutan Peralatan", "Asset (Contra)", "Without Restriction"),
        ("2003", "Utang Bank", "Liability", "Without Restriction"),
        ("4002", "Pendapatan Bunga", "Revenue", "Without Restriction"),
        ("5004", "Beban Listrik, Air, Telepon", "Expense", "Without Restriction"),
        ("5005", "Beban Transportasi", "Expense", "Without Restriction"),
        ("5006", "Beban Perlengkapan Kantor", "Expense", "Without Restriction"),
        ("1005", "Piutang Donasi", "Asset", "Without Restriction"),
        ("2004", "Utang Usaha", "Liability", "Without Restriction"),
        ("4003", "Pendapatan Jasa", "Revenue", "Without Restriction"),
        ("5007", "Beban Gaji Program", "Expense", "Without Restriction"),
        ("5008", "Beban Penyusutan Aset", "Expense", "Without Restriction"),
        ("1006", "Investasi Jangka Panjang", "Asset", "With Restriction"),
        ("3002", "Aset Neto Terikat Temporer", "Asset Net", "With Restriction"),
        ("3003", "Aset Neto Terikat Permanen", "Asset Net", "With Restriction"),
    ]
    for acc in new_accounts:
        ws_coa.append(acc)

    # Adjust column widths
    for col in ws_coa.columns:
        max_length = 0
        column = col[0].column_letter # Get the column name
        for cell in col:
            try: # Necessary to avoid error on empty cells
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = (max_length + 2)
        ws_coa.column_dimensions[column].width = adjusted_width

    # --- Sheet 2: Jurnal Umum ---
    ws_journal = wb.create_sheet("Jurnal Umum")
    journal_headers = ["Tanggal", "Deskripsi", "No. Referensi", "Kode Akun", "Debit", "Kredit", "Aktivitas Arus Kas"]
    ws_journal.append(journal_headers)

    # Apply header style
    for col_idx, header in enumerate(journal_headers, 1):
        ws_journal.cell(row=1, column=col_idx).font = Font(bold=True)
        ws_journal.cell(row=1, column=col_idx).fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")

    # Sample Journal Entries (January - December 2025)
    journal_entries_data = []
    
    # Assuming some existing accounts for the case study
    # 1001: Kas (Asset)
    # 1002: Bank (Asset)
    # 3001: Modal Awal (Equity)
    # 4001: Pendapatan Donasi (Revenue)
    # 5001: Beban Sewa (Expense)
    # 5002: Beban Gaji (Expense)
    # 5003: Beban Operasional Lainnya (Expense)

    # January 2025
    journal_entries_data.extend([
        # Saldo Awal (assuming some initial cash and bank)
        (date(2025, 1, 1), "Saldo Awal Kas dan Bank", "SA-01-2025", "1003", 500000, 0, ""),
        (date(2025, 1, 1), "Saldo Awal Kas dan Bank", "SA-01-2025", "1004", 10000000, 0, ""),
        (date(2025, 1, 1), "Saldo Awal Kas dan Bank", "SA-01-2025", "3001", 0, 10500000, ""), # Modal Awal
        
        (date(2025, 1, 5), "Penerimaan Donasi dari Bapak Budi", "DN-01-001", "1003", 1000000, 0, "Penerimaan dari Sumbangan/Donasi"),
        (date(2025, 1, 5), "Penerimaan Donasi dari Bapak Budi", "DN-01-001", "4001", 0, 1000000, ""),
        
        (date(2025, 1, 10), "Pembayaran Sewa Kantor Bulan Januari", "SW-01-001", "5001", 2000000, 0, ""),
        (date(2025, 1, 10), "Pembayaran Sewa Kantor Bulan Januari", "SW-01-001", "1004", 0, 2000000, "Pembayaran Kas untuk Beban Administrasi"),

        (date(2025, 1, 15), "Pembelian Perlengkapan Kantor", "PL-01-001", "5006", 300000, 0, ""),
        (date(2025, 1, 15), "Pembelian Perlengkapan Kantor", "PL-01-001", "1003", 0, 300000, "Pembayaran Kas untuk Beban Administrasi"),
    ])

    # February 2025
    journal_entries_data.extend([
        (date(2025, 2, 1), "Penerimaan Donasi dari Ibu Siti", "DN-02-001", "1004", 2500000, 0, "Penerimaan dari Sumbangan/Donasi"),
        (date(2025, 2, 1), "Penerimaan Donasi dari Ibu Siti", "DN-02-001", "4001", 0, 2500000, ""),

        (date(2025, 2, 10), "Pembayaran Gaji Karyawan Februari", "GJ-02-001", "5002", 4000000, 0, ""),
        (date(2025, 2, 10), "Pembayaran Gaji Karyawan Februari", "GJ-02-001", "1004", 0, 4000000, "Pembayaran Kas untuk Gaji dan Karyawan"),

        (date(2025, 2, 20), "Pembayaran Tagihan Listrik & Air", "TL-02-001", "5004", 500000, 0, ""),
        (date(2025, 2, 20), "Pembayaran Tagihan Listrik & Air", "TL-02-001", "1004", 0, 500000, "Pembayaran Kas untuk Beban Administrasi"),
    ])

    # March 2025
    journal_entries_data.extend([
        (date(2025, 3, 5), "Penerimaan Donasi Terikat untuk Program A", "DN-03-001", "1004", 5000000, 0, "Penerimaan dari Sumbangan/Donasi"),
        (date(2025, 3, 5), "Penerimaan Donasi Terikat untuk Program A", "DN-03-001", "3002", 0, 5000000, ""), # Aset Neto Terikat Temporer

        (date(2025, 3, 15), "Pembayaran Biaya Program A", "PR-03-001", "5007", 3000000, 0, ""), # Beban Gaji Program
        (date(2025, 3, 15), "Pembayaran Biaya Program A", "PR-03-001", "1004", 0, 3000000, "Pembayaran Kas untuk Beban Program"),
    ])

    # April 2025
    journal_entries_data.extend([
        (date(2025, 4, 1), "Penerimaan Jasa Konsultasi", "JS-04-001", "1004", 1500000, 0, "Penerimaan dari Jasa Layanan/Unit Bisnis"),
        (date(2025, 4, 1), "Penerimaan Jasa Konsultasi", "JS-04-001", "4003", 0, 1500000, ""), # Pendapatan Jasa

        (date(2025, 4, 10), "Pembelian Peralatan Kantor", "AT-04-001", "1501", 1000000, 0, ""), # Assuming 1501 is Peralatan Kantor (Asset)
        (date(2025, 4, 10), "Pembelian Peralatan Kantor", "AT-04-001", "1004", 0, 1000000, "Pembelian Peralatan Kantor Baru"),
    ])

    # May 2025
    journal_entries_data.extend([
        (date(2025, 5, 5), "Penerimaan Donasi Umum", "DN-05-001", "1004", 1000000, 0, "Penerimaan dari Sumbangan/Donasi"),
        (date(2025, 5, 5), "Penerimaan Donasi Umum", "DN-05-001", "4001", 0, 1000000, ""),

        (date(2025, 5, 15), "Pembayaran Utang Usaha", "UU-05-001", "2004", 750000, 0, ""), # Utang Usaha
        (date(2025, 5, 15), "Pembayaran Utang Usaha", "UU-05-001", "1004", 0, 750000, "Pembayaran Kas untuk Beban Administrasi"),
    ])

    # June 2025
    journal_entries_data.extend([
        (date(2025, 6, 1), "Penerimaan Bunga Bank", "BN-06-001", "1004", 50000, 0, ""),
        (date(2025, 6, 1), "Penerimaan Bunga Bank", "BN-06-001", "4002", 0, 50000, ""), # Pendapatan Bunga

        (date(2025, 6, 10), "Pembayaran Transportasi Kegiatan", "TR-06-001", "5005", 200000, 0, ""), # Beban Transportasi
        (date(2025, 6, 10), "Pembayaran Transportasi Kegiatan", "TR-06-001", "1003", 0, 200000, "Pembayaran Kas untuk Beban Program"),
    ])

    # July 2025
    journal_entries_data.extend([
        (date(2025, 7, 1), "Penerimaan Donasi Terikat untuk Aset", "DN-07-001", "1004", 7000000, 0, "Penerimaan Sumbangan yang Dibatasi untuk Bangunan"),
        (date(2025, 7, 1), "Penerimaan Donasi Terikat untuk Aset", "DN-07-001", "3003", 0, 7000000, ""), # Aset Neto Terikat Permanen

        (date(2025, 7, 15), "Pembelian Investasi Jangka Panjang", "INV-07-001", "1006", 6000000, 0, ""), # Investasi Jangka Panjang
        (date(2025, 7, 15), "Pembelian Investasi Jangka Panjang", "INV-07-001", "1004", 0, 6000000, "Perolehan Investasi Jangka Panjang"),
    ])

    # August 2025
    journal_entries_data.extend([
        (date(2025, 8, 5), "Penerimaan Donasi Umum", "DN-08-001", "1004", 1200000, 0, "Penerimaan dari Sumbangan/Donasi"),
        (date(2025, 8, 5), "Penerimaan Donasi Umum", "DN-08-001", "4001", 0, 1200000, ""),

        (date(2025, 8, 10), "Pembayaran Gaji Karyawan Agustus", "GJ-08-001", "5002", 4000000, 0, ""),
        (date(2025, 8, 10), "Pembayaran Gaji Karyawan Agustus", "GJ-08-001", "1004", 0, 4000000, "Pembayaran Kas untuk Gaji dan Karyawan"),
    ])

    # September 2025
    journal_entries_data.extend([
        (date(2025, 9, 1), "Penerimaan Jasa Pelatihan", "JS-09-001", "1004", 2000000, 0, "Penerimaan dari Jasa Layanan/Unit Bisnis"),
        (date(2025, 9, 1), "Penerimaan Jasa Pelatihan", "JS-09-001", "4003", 0, 2000000, ""),

        (date(2025, 9, 15), "Pembayaran Biaya Operasional Lainnya", "OP-09-001", "5003", 800000, 0, ""),
        (date(2025, 9, 15), "Pembayaran Biaya Operasional Lainnya", "OP-09-001", "1004", 0, 800000, "Pembayaran Kas untuk Beban Administrasi"),
    ])

    # October 2025
    journal_entries_data.extend([
        (date(2025, 10, 5), "Penerimaan Donasi Terikat untuk Program B", "DN-10-001", "1004", 3000000, 0, "Penerimaan dari Sumbangan/Donasi"),
        (date(2025, 10, 5), "Penerimaan Donasi Terikat untuk Program B", "DN-10-001", "3002", 0, 3000000, ""),

        (date(2025, 10, 10), "Pembayaran Biaya Program B", "PR-10-001", "5007", 2500000, 0, ""),
        (date(2025, 10, 10), "Pembayaran Biaya Program B", "PR-10-001", "1004", 0, 2500000, "Pembayaran Kas untuk Beban Program"),
    ])

    # November 2025
    journal_entries_data.extend([
        (date(2025, 11, 1), "Penerimaan Donasi Umum", "DN-11-001", "1004", 1500000, 0, "Penerimaan dari Sumbangan/Donasi"),
        (date(2025, 11, 1), "Penerimaan Donasi Umum", "DN-11-001", "4001", 0, 1500000, ""),

        (date(2025, 11, 15), "Pembayaran Utang Bank", "UB-11-001", "2003", 1000000, 0, ""), # Utang Bank
        (date(2025, 11, 15), "Pembayaran Utang Bank", "UB-11-001", "1004", 0, 1000000, "Pembayaran Pokok Pinjaman"),
    ])

    # December 2025
    journal_entries_data.extend([
        (date(2025, 12, 1), "Penyusutan Peralatan Kantor", "PS-12-001", "5008", 100000, 0, ""), # Beban Penyusutan Aset
        (date(2025, 12, 1), "Penyusutan Peralatan Kantor", "PS-12-001", "1502", 0, 100000, ""), # Assuming 1502 is Akumulasi Penyusutan Peralatan (Contra Asset)

        (date(2025, 12, 10), "Penerimaan Donasi Akhir Tahun", "DN-12-001", "1004", 3000000, 0, "Penerimaan dari Sumbangan/Donasi"),
        (date(2025, 12, 10), "Penerimaan Donasi Akhir Tahun", "DN-12-001", "4001", 0, 3000000, ""),
    ])

    for entry in journal_entries_data:
        ws_journal.append(entry)

    # Adjust column widths for journal sheet
    for col in ws_journal.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = (max_length + 2)
        ws_journal.column_dimensions[column].width = adjusted_width

    wb.save(file_path)
    print(f"Studi kasus Excel berhasil dibuat di: {file_path}")

if __name__ == "__main__":
    # Ensure the root directory is correct
    root_dir = "/home/tomychaky/aplikasi-yayasan-isak35"
    output_file = os.path.join(root_dir, "studi_kasus_yayasan_2025.xlsx")
    create_case_study_excel(output_file)
