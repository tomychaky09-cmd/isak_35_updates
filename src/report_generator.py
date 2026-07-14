import os
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

# Gunakan try-except untuk berjaga-jaga jika file pdf_report_lab.py belum sempurna
try:
    from src.pdf_report_lab import AnnualReportPDF
except ImportError:
    AnnualReportPDF = None

class ReportGenerator:
    def __init__(self, db_manager):
        self.db = db_manager

    # --- FUNGSI BANTUAN ---
    def _is_match(self, value, targets):
        """Membantu pencarian kata kunci yang lebih toleran terhadap huruf besar/kecil dan spasi"""
        val = str(value or '').lower().strip()
        return any(t.lower() in val for t in targets)

    # --- LAPORAN KEUANGAN ---
    def get_trial_balance_report(self):
        data = self.db.get_trial_balance()
        if not data: return []
        total_debit = sum(x.get('debit', 0) for x in data)
        total_credit = sum(x.get('credit', 0) for x in data)
        data.append({'code': '', 'name': 'TOTAL', 'debit': total_debit, 'credit': total_credit})
        return data

    def get_isak35_financial_position(self):
        data = self.db.get_trial_balance()
        if not data: return {}
        
        # 1. Klasifikasi Akun
        assets = [x for x in data if self._is_match(x.get('type'), ['asset', 'aset', 'harta'])]
        contra = [x for x in data if self._is_match(x.get('type'), ['contra', 'kontra', 'akumulasi', 'penyusutan'])]
        liabs = [x for x in data if self._is_match(x.get('type'), ['liability', 'liabilitas', 'hutang', 'kewajiban'])]
        
        # 2. Total Kasar
        total_assets = sum(x.get('balance', 0) for x in assets) - sum(x.get('balance', 0) for x in contra)
        total_liabs = sum(x.get('balance', 0) for x in liabs)
        
        # 3. Aset Neto Awal (Ekuitas)
        net_assets_without = sum(x.get('balance', 0) for x in data if self._is_match(x.get('type'), ['asset net', 'aset net', 'ekuitas']) and self._is_match(x.get('category'), ['tanpa', 'without']))
        net_assets_with = sum(x.get('balance', 0) for x in data if self._is_match(x.get('type'), ['asset net', 'aset net', 'ekuitas']) and self._is_match(x.get('category'), ['dengan', 'with']))
        
        # 4. Pendapatan & Beban Berjalan
        rev_wo = sum(x.get('balance', 0) for x in data if self._is_match(x.get('type'), ['revenue', 'pendapatan']) and self._is_match(x.get('category'), ['tanpa', 'without']))
        exp_wo = sum(x.get('balance', 0) for x in data if self._is_match(x.get('type'), ['expense', 'beban']) and self._is_match(x.get('category'), ['tanpa', 'without']))
        rev_w = sum(x.get('balance', 0) for x in data if self._is_match(x.get('type'), ['revenue', 'pendapatan']) and self._is_match(x.get('category'), ['dengan', 'with']))
        exp_w = sum(x.get('balance', 0) for x in data if self._is_match(x.get('type'), ['expense', 'beban']) and self._is_match(x.get('category'), ['dengan', 'with']))
        
        # 5. Saldo Akhir
        akhir_net_without = net_assets_without + (rev_wo - exp_wo)
        akhir_net_with = net_assets_with + (rev_w - exp_w)
        total_net_assets = akhir_net_without + akhir_net_with
        
        return {
            'assets': assets + contra, 
            'total_assets': total_assets,
            'liabilities': liabs, 
            'total_liabilities': total_liabs,
            'net_assets_without': akhir_net_without,
            'net_assets_with': akhir_net_with,
            'total_net_assets': total_net_assets,
            'total_liabilities_and_net_assets': total_liabs + total_net_assets,
            'surplus_without': rev_wo - exp_wo, 
            'surplus_with': rev_w - exp_w
        }

    def get_statement_of_activities(self):
        data = self.db.get_trial_balance()
        rev = [x for x in data if self._is_match(x.get('type'), ['revenue', 'pendapatan'])]
        exp = [x for x in data if self._is_match(x.get('type'), ['expense', 'beban'])]
        tr, te = sum(x.get('balance', 0) for x in rev), sum(x.get('balance', 0) for x in exp)
        return {'revenue': rev, 'total_rev': tr, 'expenses': exp, 'total_exp': te, 'total_change': tr - te}

    def get_changes_in_net_assets_report(self):
        pos = self.get_isak35_financial_position()
        akt = self.get_statement_of_activities()
        if not pos or not akt: return []
        return [
            {"description": "Aset Neto Awal Periode", "without_restriction": pos.get('net_assets_without', 0) - pos.get('surplus_without', 0), "with_restriction": pos.get('net_assets_with', 0) - pos.get('surplus_with', 0), "total": pos.get('total_net_assets', 0) - akt.get('total_change', 0)},
            {"description": "Perubahan Periode Berjalan", "without_restriction": pos.get('surplus_without', 0), "with_restriction": pos.get('surplus_with', 0), "total": akt.get('total_change', 0)},
            {"description": "Aset Neto Akhir Periode", "without_restriction": pos.get('net_assets_without', 0), "with_restriction": pos.get('net_assets_with', 0), "total": pos.get('total_net_assets', 0)}
        ]

    def get_cash_flow_report(self):
        rows = self.db._execute_query("SELECT jd.cash_flow_activity, jd.debit, jd.credit, cfc.main_category FROM journal_details jd JOIN cash_flow_categories cfc ON jd.cash_flow_activity = cfc.name", fetch=True)
        report_data, summary = [], {}
        if rows:
            for row in rows:
                key = (row[3], row[0])
                summary[key] = summary.get(key, 0) + (float(row[2] or 0) - float(row[1] or 0))
            for cat in ["ARUS KAS DARI AKTIVITAS OPERASI", "ARUS KAS DARI AKTIVITAS INVESTASI", "ARUS KAS DARI AKTIVITAS PENDANAAN"]:
                cat_items = {k[1]: v for k, v in summary.items() if k[0] == cat}
                report_data.append({"Keterangan": cat, "Jumlah (Rp)": ""})
                for act, amt in cat_items.items(): report_data.append({"Keterangan": f"  {act}", "Jumlah (Rp)": amt})
        return {"report_data": report_data}

    # --- EKSPOR DATA ---
    def export_journals_to_excel(self, filepath):
        try:
            wb = Workbook()
            ws = wb.active
            ws.append(["Tanggal", "Deskripsi", "No Ref", "Kode Akun", "Nama Akun", "Debit", "Kredit", "Aktivitas Kas"])
            for r in self.db.get_journal_data_for_export(): 
                ws.append([r[0], r[1], r[2], r[3], r[4], r[5], r[6], r[7]])
            wb.save(filepath)
            return True
        except Exception as e:
            print(f"Excel Export Error: {e}")
            return False

    def export_all_reports_to_excel(self, filepath):
        try:
            wb = Workbook()
            
            # Helpers
            def to_dict_list(data, schema):
                if not data: return []
                return [dict(zip(schema, row)) for row in data]
                
            thin_side = Side(style='thin', color='D9D9D9')
            thin_border = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)
            
            double_bottom_side = Side(style='double', color='000000')
            thin_top_side = Side(style='thin', color='000000')
            double_bottom_border = Border(top=thin_top_side, bottom=double_bottom_side)
            
            bold_font = Font(name="Calibri", size=11, bold=True)
            regular_font = Font(name="Calibri", size=11)
            
            def apply_sheet_style(ws, title, columns):
                ws.cell(row=1, column=1, value=title).font = Font(name="Calibri", size=14, bold=True, color="1F4E78")
                ws.row_dimensions[1].height = 25
                for col_idx, col_name in enumerate(columns, 1):
                    cell = ws.cell(row=3, column=col_idx, value=col_name)
                    cell.font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
                    cell.fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
                    cell.alignment = Alignment(horizontal="center", vertical="center")
                ws.row_dimensions[3].height = 24

            def autofit_sheet(ws):
                for col in ws.columns:
                    max_len = 0
                    col_letter = get_column_letter(col[0].column)
                    for cell in col:
                        val_str = str(cell.value or '')
                        for line in val_str.split('\n'):
                            if len(line) > max_len:
                                max_len = len(line)
                    ws.column_dimensions[col_letter].width = max(max_len + 3, 12)

            # -------------------------------------------------------------
            # Sheet 1: Daftar Akun (COA)
            # -------------------------------------------------------------
            ws_coa = wb.active
            ws_coa.title = "Daftar Akun (COA)"
            apply_sheet_style(ws_coa, "DAFTAR AKUN (CHART OF ACCOUNTS)", ["Kode Akun", "Nama Akun", "Tipe Akun", "Kategori", "Catatan"])
            
            accounts_data = self.db.get_accounts()
            row_idx = 4
            for acc in accounts_data:
                ws_coa.cell(row=row_idx, column=1, value=acc[1]).alignment = Alignment(horizontal="center")
                ws_coa.cell(row=row_idx, column=2, value=acc[2])
                ws_coa.cell(row=row_idx, column=3, value=acc[3]).alignment = Alignment(horizontal="center")
                ws_coa.cell(row=row_idx, column=4, value=acc[4] or "-").alignment = Alignment(horizontal="center")
                ws_coa.cell(row=row_idx, column=5, value=acc[5] or "")
                
                for col in range(1, 6):
                    ws_coa.cell(row=row_idx, column=col).border = thin_border
                    ws_coa.cell(row=row_idx, column=col).font = regular_font
                row_idx += 1
            autofit_sheet(ws_coa)

            # -------------------------------------------------------------
            # Sheet 2: Jurnal Umum
            # -------------------------------------------------------------
            ws_jurnal = wb.create_sheet(title="Jurnal Umum")
            apply_sheet_style(ws_jurnal, "JURNAL UMUM", ["Tanggal", "No. Referensi", "Keterangan", "Kode Akun", "Nama Akun", "Debit", "Kredit", "Aktivitas Kas"])
            
            jurnal_data = self.db.get_journal_data_for_export()
            row_idx = 4
            for r in jurnal_data:
                ws_jurnal.cell(row=row_idx, column=1, value=r[0]).alignment = Alignment(horizontal="center")
                ws_jurnal.cell(row=row_idx, column=2, value=r[2]).alignment = Alignment(horizontal="center")
                ws_jurnal.cell(row=row_idx, column=3, value=r[1])
                ws_jurnal.cell(row=row_idx, column=4, value=r[3]).alignment = Alignment(horizontal="center")
                ws_jurnal.cell(row=row_idx, column=5, value=r[4])
                
                cell_deb = ws_jurnal.cell(row=row_idx, column=6, value=float(r[5] or 0))
                cell_deb.number_format = '#,##0'
                cell_deb.alignment = Alignment(horizontal="right")
                
                cell_cred = ws_jurnal.cell(row=row_idx, column=7, value=float(r[6] or 0))
                cell_cred.number_format = '#,##0'
                cell_cred.alignment = Alignment(horizontal="right")
                
                ws_jurnal.cell(row=row_idx, column=8, value=r[7] or "-").alignment = Alignment(horizontal="center")
                
                for col in range(1, 9):
                    ws_jurnal.cell(row=row_idx, column=col).border = thin_border
                    ws_jurnal.cell(row=row_idx, column=col).font = regular_font
                row_idx += 1
            autofit_sheet(ws_jurnal)

            # -------------------------------------------------------------
            # Sheet 3: Buku Besar
            # -------------------------------------------------------------
            ws_ledger = wb.create_sheet(title="Buku Besar")
            apply_sheet_style(ws_ledger, "BUKU BESAR", ["Kode Akun", "Nama Akun", "Tanggal", "No. Referensi", "Keterangan", "Debit", "Kredit", "Saldo Kumulatif"])
            
            accounts_list = to_dict_list(self.db.get_accounts(), ['id', 'code', 'name', 'type', 'category', 'notes'])
            row_idx = 4
            for acc in accounts_list:
                aid = acc['id']
                code = acc['code']
                name = acc['name']
                acc_type = str(acc['type']).lower()
                
                entries = self.db.get_ledger_entries(aid)
                running_balance = 0.0
                
                if entries:
                    for e in entries:
                        debit = float(e['debit'] or 0)
                        credit = float(e['credit'] or 0)
                        
                        if any(kw in acc_type for kw in ['asset', 'expense', 'aset', 'beban', 'harta', 'biaya']):
                            running_balance += (debit - credit)
                        else:
                            running_balance += (credit - debit)
                            
                        ws_ledger.cell(row=row_idx, column=1, value=code).alignment = Alignment(horizontal="center")
                        ws_ledger.cell(row=row_idx, column=2, value=name)
                        ws_ledger.cell(row=row_idx, column=3, value=e['date']).alignment = Alignment(horizontal="center")
                        ws_ledger.cell(row=row_idx, column=4, value=e['reference_no']).alignment = Alignment(horizontal="center")
                        ws_ledger.cell(row=row_idx, column=5, value=e['description'])
                        
                        cell_deb = ws_ledger.cell(row=row_idx, column=6, value=debit)
                        cell_deb.number_format = '#,##0'
                        cell_deb.alignment = Alignment(horizontal="right")
                        
                        cell_cred = ws_ledger.cell(row=row_idx, column=7, value=credit)
                        cell_cred.number_format = '#,##0'
                        cell_cred.alignment = Alignment(horizontal="right")
                        
                        cell_bal = ws_ledger.cell(row=row_idx, column=8, value=running_balance)
                        cell_bal.number_format = '#,##0'
                        cell_bal.alignment = Alignment(horizontal="right")
                        
                        for col in range(1, 9):
                            ws_ledger.cell(row=row_idx, column=col).border = thin_border
                            ws_ledger.cell(row=row_idx, column=col).font = regular_font
                        row_idx += 1
                else:
                    ws_ledger.cell(row=row_idx, column=1, value=code).alignment = Alignment(horizontal="center")
                    ws_ledger.cell(row=row_idx, column=2, value=name)
                    ws_ledger.cell(row=row_idx, column=3, value="-").alignment = Alignment(horizontal="center")
                    ws_ledger.cell(row=row_idx, column=4, value="-").alignment = Alignment(horizontal="center")
                    ws_ledger.cell(row=row_idx, column=5, value="Saldo Awal / Tidak ada transaksi")
                    
                    cell_deb = ws_ledger.cell(row=row_idx, column=6, value=0.0)
                    cell_deb.number_format = '#,##0'
                    cell_deb.alignment = Alignment(horizontal="right")
                    
                    cell_cred = ws_ledger.cell(row=row_idx, column=7, value=0.0)
                    cell_cred.number_format = '#,##0'
                    cell_cred.alignment = Alignment(horizontal="right")
                    
                    cell_bal = ws_ledger.cell(row=row_idx, column=8, value=0.0)
                    cell_bal.number_format = '#,##0'
                    cell_bal.alignment = Alignment(horizontal="right")
                    
                    for col in range(1, 9):
                        ws_ledger.cell(row=row_idx, column=col).border = thin_border
                        ws_ledger.cell(row=row_idx, column=col).font = regular_font
                    row_idx += 1
            autofit_sheet(ws_ledger)

            # -------------------------------------------------------------
            # Sheet 4: Neraca Saldo
            # -------------------------------------------------------------
            ws_tb = wb.create_sheet(title="Neraca Saldo")
            apply_sheet_style(ws_tb, "NERACA SALDO", ["Kode Akun", "Nama Akun", "Debit", "Kredit"])
            
            tb_data = self.get_trial_balance_report()
            row_idx = 4
            for i, item in enumerate(tb_data):
                is_last = (i == len(tb_data) - 1)
                
                ws_tb.cell(row=row_idx, column=1, value=item['code']).alignment = Alignment(horizontal="center")
                ws_tb.cell(row=row_idx, column=2, value=item['name'])
                
                cell_deb = ws_tb.cell(row=row_idx, column=3, value=float(item['debit'] or 0))
                cell_deb.number_format = '#,##0'
                cell_deb.alignment = Alignment(horizontal="right")
                
                cell_cred = ws_tb.cell(row=row_idx, column=4, value=float(item['credit'] or 0))
                cell_cred.number_format = '#,##0'
                cell_cred.alignment = Alignment(horizontal="right")
                
                if is_last:
                    for col in range(1, 5):
                        ws_tb.cell(row=row_idx, column=col).font = bold_font
                        ws_tb.cell(row=row_idx, column=col).border = double_bottom_border
                else:
                    for col in range(1, 5):
                        ws_tb.cell(row=row_idx, column=col).font = regular_font
                        ws_tb.cell(row=row_idx, column=col).border = thin_border
                row_idx += 1
            autofit_sheet(ws_tb)

            # -------------------------------------------------------------
            # Sheet 5: Posisi Keuangan (Neraca)
            # -------------------------------------------------------------
            ws_pos = wb.create_sheet(title="Posisi Keuangan")
            apply_sheet_style(ws_pos, "LAPORAN POSISI KEUANGAN", ["Keterangan", "Jumlah (Rp)"])
            
            pos_data = self.get_isak35_financial_position()
            row_idx = 4
            
            def write_section_header(ws, row, title, color):
                ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=2)
                cell = ws.cell(row=row, column=1, value=title)
                cell.font = bold_font
                cell.fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
                ws.row_dimensions[row].height = 20

            # Section ASET
            write_section_header(ws_pos, row_idx, "ASET", "DDEBF7")
            row_idx += 1
            for a in pos_data.get('assets', []):
                ws_pos.cell(row=row_idx, column=1, value=a['name'])
                cell_val = ws_pos.cell(row=row_idx, column=2, value=float(a['balance'] or 0))
                cell_val.number_format = '#,##0'
                cell_val.alignment = Alignment(horizontal="right")
                ws_pos.cell(row=row_idx, column=1).border = thin_border
                ws_pos.cell(row=row_idx, column=2).border = thin_border
                row_idx += 1
            
            # Total Aset
            ws_pos.cell(row=row_idx, column=1, value="TOTAL ASET").font = bold_font
            cell_tot_assets = ws_pos.cell(row=row_idx, column=2, value=float(pos_data.get('total_assets', 0)))
            cell_tot_assets.font = bold_font
            cell_tot_assets.number_format = '#,##0'
            cell_tot_assets.alignment = Alignment(horizontal="right")
            ws_pos.cell(row=row_idx, column=1).border = double_bottom_border
            ws_pos.cell(row=row_idx, column=2).border = double_bottom_border
            row_idx += 2

            # Section LIABILITAS
            write_section_header(ws_pos, row_idx, "LIABILITAS", "F8CBAD")
            row_idx += 1
            for l in pos_data.get('liabilities', []):
                ws_pos.cell(row=row_idx, column=1, value=l['name'])
                cell_val = ws_pos.cell(row=row_idx, column=2, value=float(l['balance'] or 0))
                cell_val.number_format = '#,##0'
                cell_val.alignment = Alignment(horizontal="right")
                ws_pos.cell(row=row_idx, column=1).border = thin_border
                ws_pos.cell(row=row_idx, column=2).border = thin_border
                row_idx += 1
            
            # Total Liabilitas
            ws_pos.cell(row=row_idx, column=1, value="TOTAL LIABILITAS").font = bold_font
            cell_tot_liab = ws_pos.cell(row=row_idx, column=2, value=float(pos_data.get('total_liabilities', 0)))
            cell_tot_liab.font = bold_font
            cell_tot_liab.number_format = '#,##0'
            cell_tot_liab.alignment = Alignment(horizontal="right")
            ws_pos.cell(row=row_idx, column=1).border = double_bottom_border
            ws_pos.cell(row=row_idx, column=2).border = double_bottom_border
            row_idx += 2

            # Section ASET NETO
            write_section_header(ws_pos, row_idx, "ASET NETO", "E2EFDA")
            row_idx += 1
            
            ws_pos.cell(row=row_idx, column=1, value="Tanpa Pembatasan")
            cell_val1 = ws_pos.cell(row=row_idx, column=2, value=float(pos_data.get('net_assets_without', 0)))
            cell_val1.number_format = '#,##0'
            cell_val1.alignment = Alignment(horizontal="right")
            ws_pos.cell(row=row_idx, column=1).border = thin_border
            ws_pos.cell(row=row_idx, column=2).border = thin_border
            row_idx += 1
            
            ws_pos.cell(row=row_idx, column=1, value="Dengan Pembatasan")
            cell_val2 = ws_pos.cell(row=row_idx, column=2, value=float(pos_data.get('net_assets_with', 0)))
            cell_val2.number_format = '#,##0'
            cell_val2.alignment = Alignment(horizontal="right")
            ws_pos.cell(row=row_idx, column=1).border = thin_border
            ws_pos.cell(row=row_idx, column=2).border = thin_border
            row_idx += 1
            
            # Total Aset Neto
            ws_pos.cell(row=row_idx, column=1, value="TOTAL ASET NETO").font = bold_font
            cell_tot_net = ws_pos.cell(row=row_idx, column=2, value=float(pos_data.get('total_net_assets', 0)))
            cell_tot_net.font = bold_font
            cell_tot_net.number_format = '#,##0'
            cell_tot_net.alignment = Alignment(horizontal="right")
            ws_pos.cell(row=row_idx, column=1).border = double_bottom_border
            ws_pos.cell(row=row_idx, column=2).border = double_bottom_border
            row_idx += 2

            # Total Liabilitas dan Aset Neto
            ws_pos.cell(row=row_idx, column=1, value="TOTAL LIABILITAS DAN ASET NETO").font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
            ws_pos.cell(row=row_idx, column=1).fill = PatternFill(start_color="262626", end_color="262626", fill_type="solid")
            
            cell_tot_all = ws_pos.cell(row=row_idx, column=2, value=float(pos_data.get('total_liabilities_and_net_assets', 0)))
            cell_tot_all.font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
            cell_tot_all.fill = PatternFill(start_color="262626", end_color="262626", fill_type="solid")
            cell_tot_all.number_format = '#,##0'
            cell_tot_all.alignment = Alignment(horizontal="right")
            ws_pos.cell(row=row_idx, column=1).border = double_bottom_border
            ws_pos.cell(row=row_idx, column=2).border = double_bottom_border
            autofit_sheet(ws_pos)

            # -------------------------------------------------------------
            # Sheet 6: Laporan Aktivitas
            # -------------------------------------------------------------
            ws_act = wb.create_sheet(title="Laporan Aktivitas")
            apply_sheet_style(ws_act, "LAPORAN AKTIVITAS", ["Keterangan", "Jumlah (Rp)"])
            
            act_data = self.get_statement_of_activities()
            row_idx = 4
            
            # Section PENDAPATAN
            write_section_header(ws_act, row_idx, "PENDAPATAN", "DDEBF7")
            row_idx += 1
            for r in act_data.get('revenue', []):
                ws_act.cell(row=row_idx, column=1, value=r['name'])
                cell_val = ws_act.cell(row=row_idx, column=2, value=float(r['balance'] or 0))
                cell_val.number_format = '#,##0'
                cell_val.alignment = Alignment(horizontal="right")
                ws_act.cell(row=row_idx, column=1).border = thin_border
                ws_act.cell(row=row_idx, column=2).border = thin_border
                row_idx += 1
            
            # Total Pendapatan
            ws_act.cell(row=row_idx, column=1, value="TOTAL PENDAPATAN").font = bold_font
            cell_tot_rev = ws_act.cell(row=row_idx, column=2, value=float(act_data.get('total_rev', 0)))
            cell_tot_rev.font = bold_font
            cell_tot_rev.number_format = '#,##0'
            cell_tot_rev.alignment = Alignment(horizontal="right")
            ws_act.cell(row=row_idx, column=1).border = double_bottom_border
            ws_act.cell(row=row_idx, column=2).border = double_bottom_border
            row_idx += 2

            # Section BEBAN
            write_section_header(ws_act, row_idx, "BEBAN", "FFF2CC")
            row_idx += 1
            for e in act_data.get('expenses', []):
                ws_act.cell(row=row_idx, column=1, value=e['name'])
                cell_val = ws_act.cell(row=row_idx, column=2, value=float(e['balance'] or 0))
                cell_val.number_format = '#,##0'
                cell_val.alignment = Alignment(horizontal="right")
                ws_act.cell(row=row_idx, column=1).border = thin_border
                ws_act.cell(row=row_idx, column=2).border = thin_border
                row_idx += 1
            
            # Total Beban
            ws_act.cell(row=row_idx, column=1, value="TOTAL BEBAN").font = bold_font
            cell_tot_exp = ws_act.cell(row=row_idx, column=2, value=float(act_data.get('total_exp', 0)))
            cell_tot_exp.font = bold_font
            cell_tot_exp.number_format = '#,##0'
            cell_tot_exp.alignment = Alignment(horizontal="right")
            ws_act.cell(row=row_idx, column=1).border = double_bottom_border
            ws_act.cell(row=row_idx, column=2).border = double_bottom_border
            row_idx += 2

            # Perubahan Aset Neto
            ws_act.cell(row=row_idx, column=1, value="PERUBAHAN ASET NETO (SURPLUS/DEFISIT)").font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
            ws_act.cell(row=row_idx, column=1).fill = PatternFill(start_color="262626", end_color="262626", fill_type="solid")
            
            cell_change = ws_act.cell(row=row_idx, column=2, value=float(act_data.get('total_change', 0)))
            cell_change.font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
            cell_change.fill = PatternFill(start_color="262626", end_color="262626", fill_type="solid")
            cell_change.number_format = '#,##0'
            cell_change.alignment = Alignment(horizontal="right")
            ws_act.cell(row=row_idx, column=1).border = double_bottom_border
            ws_act.cell(row=row_idx, column=2).border = double_bottom_border
            autofit_sheet(ws_act)

            # -------------------------------------------------------------
            # Sheet 7: Laporan Perubahan Aset Neto
            # -------------------------------------------------------------
            ws_net = wb.create_sheet(title="Perubahan Aset Neto")
            apply_sheet_style(ws_net, "LAPORAN PERUBAHAN ASET NETO", ["Deskripsi", "Tanpa Pembatasan (Rp)", "Dengan Pembatasan (Rp)", "Total (Rp)"])
            
            net_data = self.get_changes_in_net_assets_report()
            row_idx = 4
            for i, row in enumerate(net_data):
                is_last = (i == len(net_data) - 1)
                ws_net.cell(row=row_idx, column=1, value=row['description'])
                
                val_wo = ws_net.cell(row=row_idx, column=2, value=float(row['without_restriction']))
                val_wo.number_format = '#,##0'
                val_wo.alignment = Alignment(horizontal="right")
                
                val_w = ws_net.cell(row=row_idx, column=3, value=float(row['with_restriction']))
                val_w.number_format = '#,##0'
                val_w.alignment = Alignment(horizontal="right")
                
                val_tot = ws_net.cell(row=row_idx, column=4, value=float(row['total']))
                val_tot.number_format = '#,##0'
                val_tot.alignment = Alignment(horizontal="right")
                
                if is_last:
                    for col in range(1, 5):
                        ws_net.cell(row=row_idx, column=col).font = bold_font
                        ws_net.cell(row=row_idx, column=col).border = double_bottom_border
                else:
                    for col in range(1, 5):
                        ws_net.cell(row=row_idx, column=col).font = regular_font
                        ws_net.cell(row=row_idx, column=col).border = thin_border
                row_idx += 1
            autofit_sheet(ws_net)

            # -------------------------------------------------------------
            # Sheet 8: Laporan Arus Kas
            # -------------------------------------------------------------
            ws_cf = wb.create_sheet(title="Arus Kas")
            apply_sheet_style(ws_cf, "LAPORAN ARUS KAS", ["Keterangan", "Jumlah (Rp)"])
            
            cf_data = self.get_cash_flow_report().get('report_data', [])
            row_idx = 4
            net_cash_flow = 0.0
            for row in cf_data:
                ket = row.get('Keterangan', '').strip()
                val_raw = row.get('Jumlah (Rp)', '')
                
                if val_raw == "":
                    ws_cf.merge_cells(start_row=row_idx, start_column=1, end_row=row_idx, end_column=2)
                    cell = ws_cf.cell(row=row_idx, column=1, value=ket)
                    cell.font = bold_font
                    cell.fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
                    ws_cf.row_dimensions[row_idx].height = 20
                    row_idx += 1
                else:
                    val = float(val_raw or 0)
                    net_cash_flow += val
                    
                    ws_cf.cell(row=row_idx, column=1, value="  " + ket)
                    cell_val = ws_cf.cell(row=row_idx, column=2, value=val)
                    cell_val.number_format = '#,##0'
                    cell_val.alignment = Alignment(horizontal="right")
                    ws_cf.cell(row=row_idx, column=1).border = thin_border
                    ws_cf.cell(row=row_idx, column=2).border = thin_border
                    row_idx += 1
            
            row_idx += 1
            ws_cf.cell(row=row_idx, column=1, value="KENAIKAN / (PENURUNAN) BERSIH KAS").font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
            ws_cf.cell(row=row_idx, column=1).fill = PatternFill(start_color="262626", end_color="262626", fill_type="solid")
            
            cell_tot = ws_cf.cell(row=row_idx, column=2, value=net_cash_flow)
            cell_tot.font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
            cell_tot.fill = PatternFill(start_color="262626", end_color="262626", fill_type="solid")
            cell_tot.number_format = '#,##0'
            cell_tot.alignment = Alignment(horizontal="right")
            ws_cf.cell(row=row_idx, column=1).border = double_bottom_border
            ws_cf.cell(row=row_idx, column=2).border = double_bottom_border
            autofit_sheet(ws_cf)

            # -------------------------------------------------------------
            # Sheet 9: Catatan Atas Laporan Keuangan (CALK)
            # -------------------------------------------------------------
            ws_calk = wb.create_sheet(title="Catatan Laporan (CALK)")
            apply_sheet_style(ws_calk, "CATATAN ATAS LAPORAN KEUANGAN (CALK)", ["Bab / Seksi", "Penjelasan / Catatan"])
            
            calk_notes = self.db.get_calk_notes()
            row_idx = 4
            for note in calk_notes:
                ws_calk.cell(row=row_idx, column=1, value=note[1]).font = bold_font
                ws_calk.cell(row=row_idx, column=1).alignment = Alignment(vertical="top")
                
                cell_content = ws_calk.cell(row=row_idx, column=2, value=note[2])
                cell_content.alignment = Alignment(wrap_text=True, vertical="top")
                
                ws_calk.cell(row=row_idx, column=1).border = thin_border
                ws_calk.cell(row=row_idx, column=2).border = thin_border
                row_idx += 1
            
            ws_calk.column_dimensions['A'].width = 30
            ws_calk.column_dimensions['B'].width = 80

            wb.save(filepath)
            return True
        except Exception as e:
            print(f"Excel Export Error: {e}")
            return False

    def export_annual_report_to_pdf(self, filepath, year):
        if AnnualReportPDF:
            try:
                profile = self.db.get_foundation_profile()
                settings = self.db.get_annual_report_settings(year) or {}
                
                # Format profile data
                prof_data = {
                    'name': profile.get('name', 'Yayasan ISAK 35'),
                    'address': str(profile.get('address') or '').lstrip('-').strip(),
                    'year': year,
                    'vision': settings.get('vision', ''),
                    'mission': settings.get('mission', '')
                }
                
                # Logika Struktur Organisasi Fleksibel (Parsing dari settings atau fallback ke profile)
                org_structure_raw = settings.get('organizational_structure', '')
                organization = []
                
                if org_structure_raw and ':' in org_structure_raw:
                    lines = [l.strip() for l in org_structure_raw.replace('\r\n', '\n').split('\n') if ':' in l]
                    for line in lines:
                        parts = line.split(':', 1)
                        organization.append({
                            "position": parts[0].strip(), 
                            "name": parts[1].strip().lstrip('-').strip()
                        })
                
                if not organization:
                    organization = [
                        {"position": "Ketua Pembina", "name": str(profile.get('pembina_name') or '').lstrip('-').strip() or '...................'},
                        {"position": "Ketua Pengawas", "name": str(profile.get('pengawas_name') or '').lstrip('-').strip() or '...................'},
                        {"position": "Ketua Pengurus", "name": str(profile.get('leader_name') or '').lstrip('-').strip() or '...................'},
                        {"position": "Bendahara", "name": "..................."}
                    ]
                
                # Compile complete data dictionary for PDF report
                pdf_data = {
                    'profile': prof_data,
                    'organization': organization,
                    'performance': settings.get('program_summary', ''),
                    'program_edu': settings.get('program_detail_edu', ''),
                    'program_social': settings.get('program_detail_social', ''),
                    'evaluation': settings.get('evaluation_constraints', ''),
                    'future_plans': settings.get('future_plans', ''),
                    'financials': {
                        'position': self.get_isak35_financial_position(),
                        'activities': self.get_statement_of_activities(),
                        'changes_net_assets': self.get_changes_in_net_assets_report(),
                        'cash_flow': self.get_cash_flow_report().get('report_data', [])
                    },
                    'calk': [{'id': n[0], 'section_title': n[1], 'content': n[2]} for n in self.db.get_calk_notes()]
                }
                
                pdf = AnnualReportPDF(pdf_data)
                pdf.generate(filepath)
                return True
            except Exception as e:
                print(f"PDF Export Error: {e}")
                return False
        return False