from flask import Flask, render_template, jsonify, request, session, redirect, url_for, flash, send_file
from src.database_manager import DatabaseManager
from src.report_generator import ReportGenerator
import os
import io
from functools import wraps
from datetime import datetime
import tempfile
from openpyxl import load_workbook, Workbook

app = Flask(__name__, 
            template_folder='templates',
            static_folder='static')
app.secret_key = 'isak35_secret_key'

# --- BASIC HELPERS & DECORATORS ---
def login_required(f):
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if 'user' not in session: return redirect(url_for('login'))
        return f(*args, **kwargs)
    return decorated_function

def to_dict_list(data, schema):
    if not data: return []
    if isinstance(data, list) and len(data) > 0 and isinstance(data[0], dict):
        return data
    return [dict(zip(schema, row)) for row in data]

# --- HELPERS UNTUK EXCEL TEMPLATE ---
def create_excel_template(filename, headers, sheet_name="Template"):
    output = io.BytesIO()
    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name
    ws.append(headers)
    # Atur lebar kolom sedikit lebih luas
    for i in range(1, len(headers) + 1):
        ws.column_dimensions[chr(64 + i)].width = 20
    wb.save(output)
    output.seek(0)
    return send_file(output, as_attachment=True, download_name=filename, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')

@app.route('/download-template/journal')
@login_required
def download_template_journal():
    return create_excel_template('template_jurnal.xlsx', ['Tanggal (YYYY-MM-DD)', 'No Ref', 'Keterangan', 'Kode Akun', 'Debit', 'Kredit', 'Arus Kas (Opsional)'], "Jurnal")

# --- ROUTES TEMPLATE ---
@app.route('/download-template/coa')
@login_required
def download_template_coa():
    return create_excel_template('template_coa.xlsx', ['Kode', 'Nama Akun', 'Tipe', 'Kategori', 'Catatan'], "COA")

@app.route('/download-template/cf-cats')
@login_required
def download_template_cf_cats():
    return create_excel_template('template_kategori_arus_kas.xlsx', ['Nama Kategori', 'Kategori Utama'], "CashFlowCats")

# Inisialisasi Database
base_dir = os.path.dirname(os.path.abspath(__file__))
db_path = os.path.join(base_dir, 'database', 'foundation_finance.db')
if not os.path.exists(os.path.dirname(db_path)):
    os.makedirs(os.path.dirname(db_path))

db = DatabaseManager(db_path)
rg = ReportGenerator(db)
SYNC_TOKEN = "ISAK35_Qudwah_Sync_2026"

@app.context_processor
def inject_db_info():
    return dict(db_type=db.db_type, now=datetime.now().strftime('%Y-%m-%d'))

@app.route('/login', methods=['GET', 'POST'])
def login():
    if request.method == 'POST':
        u = request.form.get('username'); p = request.form.get('password')
        user = db.verify_login(u, p)
        if user: session['user'] = user; return redirect(url_for('dashboard'))
        flash('Username/Password salah!', 'danger')
    return render_template('login.html')

@app.route('/logout')
def logout():
    session.pop('user', None); return redirect(url_for('login'))

@app.route('/')
@login_required
def dashboard():
    profile = db.get_foundation_profile()
    try:
        accounts = to_dict_list(db.get_accounts(), ['id', 'code', 'name', 'type', 'category', 'notes'])
        journals = to_dict_list(db.get_journal_summaries(), ['id', 'date', 'description', 'reference_no'])
        donors = to_dict_list(db.get_donors(), ['id', 'name', 'phone', 'address', 'type', 'description'])
        tb = db.get_trial_balance()
        
        asset_types = ['asset', 'aset', 'harta']
        contra_types = ['contra', 'kontra', 'akumulasi', 'penyusutan']
        revenue_types = ['revenue', 'pendapatan', 'penerimaan']
        expense_types = ['expense', 'beban', 'biaya', 'pengeluaran']
        
        def is_type_exact(actual, target_list):
            if not actual: return False
            return str(actual).lower().strip() in [t.lower() for t in target_list]
        
        def is_type_contains(actual, target_list):
            if not actual: return False
            return any(t in str(actual).lower().strip() for t in target_list)

        total_asset = sum(x.get('balance', 0) for x in tb if is_type_exact(x.get('type'), asset_types))
        total_contra = sum(x.get('balance', 0) for x in tb if is_type_contains(x.get('type'), contra_types))
        
        total_rev = sum(x.get('balance', 0) for x in tb if is_type_exact(x.get('type'), revenue_types))
        total_exp = sum(x.get('balance', 0) for x in tb if is_type_exact(x.get('type'), expense_types))
        
        return render_template('dashboard.html', 
                               profile=profile, 
                               total_accounts=len(accounts), 
                               total_journals=len(journals), 
                               total_donors=len(donors), 
                               total_asset=total_asset - total_contra, 
                               total_revenue=total_rev, 
                               total_expense=total_exp, 
                               surplus=total_rev - total_exp)
    except Exception as e:
        print(f"Dashboard Error: {e}")
        return render_template('dashboard.html', profile=profile, total_accounts=0, total_journals=0, total_donors=0, total_asset=0, total_revenue=0, total_expense=0, surplus=0)

# --- COA ---
@app.route('/coa')
@login_required
def coa_page():
    accounts = to_dict_list(db.get_accounts(), ['id', 'code', 'name', 'type', 'category', 'notes'])
    return render_template('coa.html', accounts=accounts)

@app.route('/coa/add', methods=['GET', 'POST'])
@login_required
def add_coa():
    if request.method == 'POST':
        db.add_account(request.form.get('code'), request.form.get('name'), request.form.get('type'), request.form.get('category'), request.form.get('notes'))
        flash('Akun ditambahkan!', 'success'); return redirect(url_for('coa_page'))
    return render_template('coa_form.html')

@app.route('/coa/edit/<int:aid>', methods=['GET', 'POST'])
@login_required
def edit_coa(aid):
    accs = to_dict_list(db.get_accounts(), ['id', 'code', 'name', 'type', 'category', 'notes'])
    acc = next((x for x in accs if x['id'] == aid), None)
    if request.method == 'POST':
        db.update_account(aid, request.form.get('code'), request.form.get('name'), request.form.get('type'), request.form.get('category'), request.form.get('notes'))
        flash('Akun diperbarui!', 'success'); return redirect(url_for('coa_page'))
    return render_template('coa_form.html', account=acc)

@app.route('/coa/delete/<int:aid>')
@login_required
def delete_coa(aid):
    db.delete_account(aid); flash('Akun dihapus.', 'success'); return redirect(url_for('coa_page'))

@app.route('/coa/import', methods=['POST'])
@login_required
def import_coa():
    if 'file' not in request.files:
        flash('Tidak ada file dipilih', 'danger')
        return redirect(url_for('coa_page'))
    
    file = request.files['file']
    if file.filename == '':
        flash('Nama file kosong', 'danger')
        return redirect(url_for('coa_page'))

    if file and file.filename.endswith('.xlsx'):
        try:
            wb = load_workbook(file, data_only=True)
            ws = wb.active
            
            imported_count = 0
            failed_count = 0
            # Lewati header (baris 1)
            for row in ws.iter_rows(min_row=2, values_only=True):
                if not row[0]: continue # Lewati baris kosong
                try:
                    code = str(row[0])
                    name = str(row[1])
                    acc_type = str(row[2])
                    category = str(row[3]) if row[3] else ""
                    notes = str(row[4]) if row[4] else ""
                    
                    if db.add_account(code, name, acc_type, category, notes):
                        imported_count += 1
                    else:
                        failed_count += 1
                except:
                    failed_count += 1
            
            msg = f'Berhasil mengimpor {imported_count} akun!'
            if failed_count > 0: msg += f' ({failed_count} baris gagal/sudah ada)'
            flash(msg, 'success' if failed_count == 0 else 'warning')
        except Exception as e:
            flash(f'Gagal mengimpor: {str(e)}', 'danger')
    else:
        flash('Format file harus .xlsx', 'danger')
        
    return redirect(url_for('coa_page'))

# --- JOURNALS ---
@app.route('/journals')
@login_required
def journals_page():
    journals = to_dict_list(db.get_journal_summaries(), ['id', 'date', 'description', 'reference_no'])
    return render_template('journals.html', journals=journals)

@app.route('/journals/add', methods=['GET', 'POST'])
@login_required
def add_journal():
    if request.method == 'POST':
        try:
            date = request.form.get('date'); ref = request.form.get('ref_no'); desc = request.form.get('description')
            ids = request.form.getlist('account_id[]'); dbt = request.form.getlist('debit[]'); crd = request.form.getlist('credit[]'); cf = request.form.getlist('cf_activity[]')
            dets = []
            for i in range(len(ids)):
                if ids[i]: dets.append({'account_id': int(ids[i]), 'debit': float(dbt[i] or 0), 'credit': float(crd[i] or 0), 'cash_flow_activity': cf[i] if i < len(cf) else None})
            if db.add_journal_entry(date, desc, ref, dets): flash('Jurnal disimpan!', 'success'); return redirect(url_for('journals_page'))
        except Exception as e: flash(f'Error: {e}', 'danger')
    accounts = to_dict_list(db.get_accounts(), ['id', 'code', 'name', 'type', 'category', 'notes'])
    cf_cats = to_dict_list(db.get_cash_flow_categories(), ['id', 'name', 'main_category'])
    return render_template('journal_form.html', accounts=accounts, cf_cats=cf_cats)

@app.route('/journals/edit/<int:jid>', methods=['GET', 'POST'])
@login_required
def edit_journal(jid):
    if request.method == 'POST':
        try:
            date = request.form.get('date'); ref = request.form.get('ref_no'); desc = request.form.get('description')
            ids = request.form.getlist('account_id[]'); dbt = request.form.getlist('debit[]'); crd = request.form.getlist('credit[]'); cf = request.form.getlist('cf_activity[]')
            dets = []
            for i in range(len(ids)):
                if ids[i]: dets.append({'account_id': int(ids[i]), 'debit': float(dbt[i] or 0), 'credit': float(crd[i] or 0), 'cash_flow_activity': cf[i] if i < len(cf) else None})
            if db.update_journal_entry(jid, date, desc, ref, dets): flash('Jurnal diperbarui!', 'success'); return redirect(url_for('journals_page'))
        except Exception as e: flash(f'Error: {e}', 'danger')
    
    raw_data = db.get_journal_details(jid)
    if not raw_data: return "Jurnal tidak ditemukan", 404
    
    data = {
        'header': dict(zip(['date', 'description', 'reference_no'], raw_data['header'])),
        'details': to_dict_list(raw_data['details'], ['code', 'name', 'debit', 'credit', 'cash_flow_activity', 'account_id'])
    }
    accounts = to_dict_list(db.get_accounts(), ['id', 'code', 'name', 'type', 'category', 'notes'])
    cf_cats = to_dict_list(db.get_cash_flow_categories(), ['id', 'name', 'main_category'])
    return render_template('journal_form.html', data=data, journal_id=jid, accounts=accounts, cf_cats=cf_cats)

@app.route('/journals/<int:jid>')
@login_required
def journal_detail(jid):
    raw_data = db.get_journal_details(jid)
    if not raw_data: return "Bukan Jurnal", 404
    data = {
        'header': dict(zip(['date', 'description', 'reference_no'], raw_data['header'])),
        'details': to_dict_list(raw_data['details'], ['code', 'name', 'debit', 'credit', 'cash_flow_activity'])
    }
    return render_template('journal_detail.html', data=data)

@app.route('/journals/delete/<int:jid>')
@login_required
def delete_journal(jid):
    db.delete_journal_entry(jid); flash('Jurnal dihapus.', 'success'); return redirect(url_for('journals_page'))

@app.route('/journals/export')
@login_required
def export_journals():
    with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as tmp:
        file_path = tmp.name
    if rg.export_journals_to_excel(file_path):
        return send_file(file_path, as_attachment=True, download_name=f"Daftar_Jurnal_{datetime.now().strftime('%Y%m%d')}.xlsx")
    flash("Gagal mengekspor jurnal", "danger")
    return redirect(url_for('journals_page'))

@app.route('/journals/import', methods=['POST'])
@login_required
def import_journals():
    if 'file' not in request.files:
        flash('Tidak ada file dipilih', 'danger')
        return redirect(url_for('journals_page'))
    
    file = request.files['file']
    if file.filename == '':
        flash('Nama file kosong', 'danger')
        return redirect(url_for('journals_page'))

    if file and file.filename.endswith('.xlsx'):
        try:
            wb = load_workbook(file, data_only=True)
            ws = wb.active
            
            # Map kode akun ke ID
            acc_map = {str(a['code']): a['id'] for a in to_dict_list(db.get_accounts(), ['id', 'code'])}
            
            journal_groups = {} # Key: RefNo
            for row in ws.iter_rows(min_row=2, values_only=True):
                if not row[1]: continue # Butuh RefNo
                
                date_val = str(row[0])[:10] if row[0] else datetime.now().strftime('%Y-%m-%d')
                ref = str(row[1])
                desc = str(row[2]) if row[2] else ""
                acc_code = str(row[3])
                debit = float(row[4] or 0)
                credit = float(row[5] or 0)
                cf = str(row[6]) if row[6] else None
                
                if acc_code not in acc_map:
                    flash(f'Gagal: Kode Akun {acc_code} tidak ditemukan!', 'danger')
                    return redirect(url_for('journals_page'))
                
                if ref not in journal_groups:
                    journal_groups[ref] = {'date': date_val, 'desc': desc, 'details': []}
                
                journal_groups[ref]['details'].append({
                    'account_id': acc_map[acc_code],
                    'debit': debit,
                    'credit': credit,
                    'cash_flow_activity': cf
                })
            
            imported_count = 0
            for ref, data in journal_groups.items():
                # Validasi Balance
                total_debit = sum(d['debit'] for d in data['details'])
                total_credit = sum(d['credit'] for d in data['details'])
                
                if abs(total_debit - total_credit) > 0.01:
                    flash(f'Gagal: Jurnal {ref} tidak balance! (D: {total_debit}, C: {total_credit})', 'danger')
                    continue
                
                if db.add_journal_entry(data['date'], data['desc'], ref, data['details']):
                    imported_count += 1
            
            flash(f'Berhasil mengimpor {imported_count} jurnal!', 'success')
        except Exception as e:
            flash(f'Gagal mengimpor: {str(e)}', 'danger')
    else:
        flash('Format file harus .xlsx', 'danger')
    return redirect(url_for('journals_page'))

# --- ASSETS ---
@app.route('/assets')
@login_required
def assets_page():
    assets = to_dict_list(db.get_assets_inventory(), ['id', 'date', 'code', 'name', 'location', 'estimated_value', 'quantity', 'description'])
    return render_template('assets.html', assets=assets)

@app.route('/assets/add', methods=['GET', 'POST'])
@login_required
def add_asset():
    if request.method == 'POST':
        db.add_asset_inventory(request.form.get('date'), request.form.get('code'), request.form.get('name'), request.form.get('location'), float(request.form.get('estimated_value') or 0), int(request.form.get('quantity') or 1), request.form.get('description'))
        flash('Aset ditambahkan!', 'success'); return redirect(url_for('assets_page'))
    return render_template('asset_form.html')

@app.route('/assets/edit/<int:aid>', methods=['GET', 'POST'])
@login_required
def edit_asset(aid):
    assets = to_dict_list(db.get_assets_inventory(), ['id', 'date', 'code', 'name', 'location', 'estimated_value', 'quantity', 'description'])
    asset = next((x for x in assets if x['id'] == aid), None)
    if request.method == 'POST':
        db.update_asset_inventory(aid, request.form.get('date'), request.form.get('code'), request.form.get('name'), request.form.get('location'), float(request.form.get('estimated_value') or 0), int(request.form.get('quantity') or 1), request.form.get('description'))
        flash('Aset diperbarui!', 'success'); return redirect(url_for('assets_page'))
    return render_template('asset_form.html', asset=asset)

@app.route('/assets/delete/<int:aid>')
@login_required
def delete_asset(aid):
    db.delete_asset_inventory(aid); flash('Aset dihapus.', 'success'); return redirect(url_for('assets_page'))

# --- DONORS ---
@app.route('/donors')
@login_required
def donors_page():
    donors = to_dict_list(db.get_donors(), ['id', 'name', 'phone', 'address', 'type', 'description'])
    return render_template('donors.html', donors=donors)

@app.route('/donors/add', methods=['GET', 'POST'])
@login_required
def add_donor():
    if request.method == 'POST':
        db.add_donor(request.form.get('name'), request.form.get('phone'), request.form.get('address'), request.form.get('type'), request.form.get('description'))
        flash('Donatur ditambahkan!', 'success'); return redirect(url_for('donors_page'))
    return render_template('donor_form.html')

@app.route('/donors/edit/<int:did>', methods=['GET', 'POST'])
@login_required
def edit_donor(did):
    donors = to_dict_list(db.get_donors(), ['id', 'name', 'phone', 'address', 'type', 'description'])
    donor = next((x for x in donors if x['id'] == did), None)
    if request.method == 'POST':
        db.update_donor(did, request.form.get('name'), request.form.get('phone'), request.form.get('address'), request.form.get('type'), request.form.get('description'))
        flash('Donatur diperbarui!', 'success'); return redirect(url_for('donors_page'))
    return render_template('donor_form.html', donor=donor)

@app.route('/donors/delete/<int:did>')
@login_required
def delete_donor(did):
    db.delete_donor(did); flash('Donatur dihapus.', 'success'); return redirect(url_for('donors_page'))

# --- LEDGER ---
@app.route('/ledger')
@login_required
def ledger_page():
    accounts = to_dict_list(db.get_accounts(), ['id', 'code', 'name', 'type', 'category', 'notes'])
    return render_template('ledger_select.html', accounts=accounts)

@app.route('/ledger/<int:aid>')
@login_required
def ledger_detail(aid):
    accs = to_dict_list(db.get_accounts(), ['id', 'code', 'name', 'type', 'category', 'notes'])
    curr = next((x for x in accs if x['id'] == aid), None)
    if not curr: return "Akun tidak ada", 404
    ents = db.get_ledger_entries(aid)
    bal = 0; res = []
    for e in ents:
        d = e['debit'] or 0; c = e['credit'] or 0
        acc_type = str(curr.get('type', '')).lower()
        if any(kw in acc_type for kw in ['asset', 'expense', 'aset', 'beban', 'harta', 'biaya']): bal += (d - c)
        else: bal += (c - d)
        e['running_balance'] = bal; res.append(e)
    return render_template('ledger_detail.html', account=curr, entries=res, final_balance=bal)

# --- KATEGORI ARUS KAS ---
@app.route('/cash-flow-cats', methods=['GET'])
@login_required
def cf_cats_page():
    cats = to_dict_list(db.get_cash_flow_categories(), ['id', 'name', 'main_category'])
    return render_template('cash_flow_cats.html', cats=cats)

@app.route('/cash-flow-cats/add', methods=['POST'])
@login_required
def add_cf_cat():
    db.add_cash_flow_category(request.form.get('name'), request.form.get('main_category'))
    flash('Kategori ditambahkan', 'success')
    return redirect(url_for('cf_cats_page'))

@app.route('/cash-flow-cats/delete/<path:name>')
@login_required
def delete_cf_cat(name):
    db.delete_cash_flow_category(name)
    flash('Kategori dihapus', 'success')
    return redirect(url_for('cf_cats_page'))

@app.route('/cash-flow-cats/import', methods=['POST'])
@login_required
def import_cf_cats():
    if 'file' not in request.files:
        flash('Tidak ada file dipilih', 'danger')
        return redirect(url_for('cf_cats_page'))
    
    file = request.files['file']
    if file.filename == '':
        flash('Nama file kosong', 'danger')
        return redirect(url_for('cf_cats_page'))

    if file and file.filename.endswith('.xlsx'):
        try:
            wb = load_workbook(file, data_only=True)
            ws = wb.active
            imported_count = 0
            failed_count = 0
            for row in ws.iter_rows(min_row=2, values_only=True):
                if not row[0]: continue
                try:
                    name = str(row[0])
                    main_cat_raw = str(row[1]).upper() if row[1] else "OPERASI"
                    
                    # Normalisasi kategori utama
                    main_cat = "ARUS KAS DARI AKTIVITAS OPERASI"
                    if "INVESTASI" in main_cat_raw: main_cat = "ARUS KAS DARI AKTIVITAS INVESTASI"
                    elif "PENDANAAN" in main_cat_raw: main_cat = "ARUS KAS DARI AKTIVITAS PENDANAAN"
                    
                    if db.add_cash_flow_category(name, main_cat):
                        imported_count += 1
                    else:
                        failed_count += 1
                except:
                    failed_count += 1
            
            msg = f'Berhasil mengimpor {imported_count} kategori!'
            if failed_count > 0: msg += f' ({failed_count} baris gagal/sudah ada)'
            flash(msg, 'success' if failed_count == 0 else 'warning')
        except Exception as e:
            flash(f'Gagal mengimpor: {str(e)}', 'danger')
    else:
        flash('Format file harus .xlsx', 'danger')
    return redirect(url_for('cf_cats_page'))

# --- REPORTS ---
@app.route('/reports/trial-balance')
@login_required
def trial_balance():
    data = rg.get_trial_balance_report()
    totals = {'debit': sum(x.get('debit', 0) for x in data if x.get('code')), 'credit': sum(x.get('credit', 0) for x in data if x.get('code'))}
    return render_template('report_trial_balance.html', data=data, totals=totals)

@app.route('/reports/activities')
@login_required
def report_activities():
    return render_template('report_activities.html', data=rg.get_statement_of_activities())

@app.route('/reports/position')
@login_required
def report_position():
    return render_template('report_position.html', data=rg.get_isak35_financial_position())

@app.route('/reports/cash-flow')
@login_required
def report_cash_flow():
    return render_template('report_cash_flow.html', data=rg.get_cash_flow_report()['report_data'])

@app.route('/settings/reset')
@login_required
def reset_data_page():
    return render_template('reset_data.html')

@app.route('/settings/reset/process', methods=['POST'])
@login_required
def reset_data_process():
    selected_tables = request.form.getlist('tables')
    if not selected_tables:
        flash("Tidak ada kategori data yang dipilih.", "warning")
        return redirect(url_for('reset_data_page'))
    
    success_count = 0
    error_msgs = []
    
    try:
        # Mapping nama dari form ke tabel DB
        table_map = {
            'journals': ['journal_details', 'journal_entries'],
            'accounts': ['accounts'],
            'donors': ['donors'],
            'assets': ['assets_inventory'],
            'cashflow': ['cash_flow_categories']
        }
        
        conn = db.get_connection()
        cursor = conn.cursor()
        
        for key in selected_tables:
            if key in table_map:
                for table in table_map[key]:
                    try:
                        cursor.execute(f"DELETE FROM {table}")
                        # Reset auto-increment jika di SQLite
                        if db.db_type == 'sqlite':
                            cursor.execute(f"DELETE FROM sqlite_sequence WHERE name='{table}'")
                        success_count += 1
                    except Exception as e:
                        error_msgs.append(f"Gagal hapus {table}: {str(e)}")
        
        conn.commit()
        conn.close()
        
        if success_count > 0:
            flash(f"Berhasil membersihkan data dari {success_count} tabel.", "success")
        if error_msgs:
            for msg in error_msgs: flash(msg, "danger")
            
    except Exception as e:
        flash(f"Error sistem: {str(e)}", "danger")
        
    return redirect(url_for('reset_data_page'))

@app.route('/settings/backup-db')
@login_required
def backup_db():
    try:
        # Gunakan path database yang sudah didefinisikan di awal (db.db_path)
        if os.path.exists(db.db_path):
            filename = f"backup_foundation_{datetime.now().strftime('%Y%m%d_%H%M')}.db"
            return send_file(db.db_path, as_attachment=True, download_name=filename)
        else:
            flash("File database tidak ditemukan!", "danger")
    except Exception as e:
        flash(f"Gagal melakukan backup: {str(e)}", "danger")
    return redirect(url_for('reset_data_page'))

@app.route('/settings/import-db', methods=['POST'])
@login_required
def import_db():
    global db, rg
    if 'file' not in request.files:
        flash('Tidak ada file dipilih', 'danger')
        return redirect(url_for('reset_data_page'))
    
    file = request.files['file']
    if file.filename == '':
        flash('Nama file kosong', 'danger')
        return redirect(url_for('reset_data_page'))

    if file and file.filename.endswith('.db'):
        try:
            # Pastikan folder database ada
            os.makedirs(os.path.dirname(db.db_path), exist_ok=True)
            
            # Ganti file database lama dengan yang baru
            file.save(db.db_path)
            
            # Re-inisialisasi koneksi database agar menggunakan file baru
            from src.database_manager import DatabaseManager
            from src.report_generator import ReportGenerator
            db = DatabaseManager(db.db_path)
            rg = ReportGenerator(db)
            
            flash('Database berhasil diimpor! Data web kini sinkron dengan data yang Anda unggah.', 'success')
        except Exception as e:
            flash(f'Gagal mengimpor database: {str(e)}', 'danger')
    else:
        flash('Format file harus .db (SQLite)', 'danger')
        
    return redirect(url_for('reset_data_page'))

@app.route('/reports/annual', methods=['GET', 'POST'])
@login_required
def annual_report_page():
    year = request.args.get('year', datetime.now().year)
    settings = db.get_annual_report_settings(year)
    return render_template('annual_report_form.html', settings=settings, year=year)

@app.route('/reports/annual/save', methods=['POST'])
@login_required
def save_annual_report():
    year = request.form.get('year')
    data = {
        'vision': request.form.get('vision'),
        'mission': request.form.get('mission'),
        'program_summary': request.form.get('program_summary'),
        'program_detail_edu': request.form.get('program_detail_edu'),
        'program_detail_social': request.form.get('program_detail_social'),
        'organizational_structure': request.form.get('organizational_structure'),
        'evaluation_constraints': request.form.get('evaluation_constraints'),
        'future_plans': request.form.get('future_plans')
    }
    if db.save_annual_report_settings(year, data):
        flash('Pengaturan laporan tahunan disimpan!', 'success')
    else:
        flash('Gagal menyimpan pengaturan!', 'danger')
    return redirect(url_for('annual_report_page', year=year))

@app.route('/reports/annual/pdf/<int:year>')
@login_required
def download_annual_report_pdf(year):
    with tempfile.NamedTemporaryFile(suffix='.pdf', delete=False) as tmp:
        file_path = tmp.name
    if rg.export_annual_report_to_pdf(file_path, year):
        return send_file(file_path, as_attachment=True, download_name=f"Laporan_Tahunan_{year}.pdf")
    flash("Gagal membuat PDF laporan tahunan", "danger")
    return redirect(url_for('annual_report_page', year=year))

@app.route('/export/excel')
@login_required
def export_excel():
    with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as tmp:
        file_path = tmp.name
    if rg.export_all_reports_to_excel(file_path):
        return send_file(file_path, as_attachment=True, download_name=f"Laporan_ISAK35_{datetime.now().strftime('%Y%m%d')}.xlsx")
    return "Gagal", 500

@app.route('/api/dashboard-stats')
@login_required
def dashboard_stats_api():
    try:
        data = db.get_journal_data_for_export()
        if not data: return jsonify({"labels": [], "income": [], "expense": []})
        stats = {}
        accs = {a['code']: a['type'] for a in to_dict_list(db.get_accounts(), ['id', 'code', 'name', 'type', 'category', 'notes'])}
        rev = ['revenue', 'pendapatan', 'penerimaan']; exp = ['expense', 'beban', 'biaya', 'pengeluaran']
        for r in data:
            m = r['Tanggal'][:7]
            if m not in stats: stats[m] = {'i': 0, 'e': 0}
            t = str(accs.get(r['Kode Akun'], '')).lower()
            if any(k in t for k in rev): stats[m]['i'] += (r['Kredit'] - r['Debit'])
            elif any(k in t for k in exp): stats[m]['e'] += (r['Debit'] - r['Kredit'])
        labels = sorted(stats.keys())[-6:]
        return jsonify({"labels": labels, "income": [stats[l]['i'] for l in labels], "expense": [stats[l]['e'] for l in labels]})
    except: return jsonify({"labels": [], "income": [], "expense": []})

# --- SYNC ---
@app.route('/api/sync/pull', methods=['GET'])
def api_pull():
    if request.headers.get('Authorization') != SYNC_TOKEN: return jsonify({"error": "No"}), 401
    return jsonify(db.generate_full_backup())

@app.route('/api/sync/push', methods=['POST'])
def api_push():
    if request.headers.get('Authorization') != SYNC_TOKEN: return jsonify({"error": "No"}), 401
    ok, msg = db.restore_full_backup(request.json)
    return jsonify({"status": "success"}) if ok else jsonify({"error": msg}), 500

if __name__ == '__main__':
    app.run(debug=True)
